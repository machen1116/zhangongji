#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
【导入】Excel → data.js（极简版）
用法：python3 excel_to_js.py <Excel文件> [data.js路径]

原理：
  1. 读取 Excel → Python dict
  2. json.dumps(dict) → 标准 JSON（同时也是合法 JS 对象字面量）
  3. 包上 const cardData = ... ; 外壳
  4. 写入 data.js
"""

import sys, os, json, shutil

try:
    import openpyxl as xl
except ImportError:
    print("❌ 请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)

# ── 常量 ─────────────────────────────────────────────
CAT_NAMES = {"law":"法律","politics":"政治","mazhe":"哲学","guanli":"管理","feifa":"非法"}
SUBJ_NAMES = {
    "xingfa":"刑法","minfa":"民法","xingzheng":"行政法",
    "xianfa":"宪法","susong":"诉讼法",
    "sixiang":"习近平新时代中国特色社会主义思想",
    "mazhe":"马克思主义基本原理",
    "guanli":"管理学","feifa":"非法",
}

# ── Excel 读取 ───────────────────────────────────────
HEADER_MAP = {
    "学科大类": "category",
    "学科编号": "subject",
    "章节名称": "chapter",
    "问题（正面）": "question",
    "答案（背面）": "answer",
    "标签": "tagText",
    "难度": "difficulty",
    "解析": "analysis",
    "考点提示": "analysis",
    "知识拓展": "expansion",
    # 兼容英文表头
    "category": "category",
    "subject": "subject",
    "chapter": "chapter",
    "question": "question",
    "answer": "answer",
    "tagText": "tagText",
    "difficulty": "difficulty",
    "analysis": "analysis",
    "expansion": "expansion",
}

def load_excel(path):
    wb = xl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("❌ Excel 中没有数据行")
        sys.exit(1)
    # 映射表头：中文 → 英文字段名
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    headers = [HEADER_MAP.get(h, h) for h in raw_headers]
    data = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h: continue
            v = row[i] if i < len(row) else None
            d[h] = str(v).strip() if v is not None else ""
        if not d.get("question") or not d.get("answer"):
            print(f"⚠️  跳过不完整行：{d}")
            continue
        data.append(d)
    return data

# ── 数据结构化 ─────────────────────────────────────
def build(data):
    struct = {}
    order = {}  # 记录章节出现顺序
    for d in data:
        cat = d["category"].strip()
        subj = d["subject"].strip()
        ch = d["chapter"].strip()
        if not cat or not subj or not ch:
            continue
        if cat not in struct:
            struct[cat] = {"name": CAT_NAMES.get(cat, cat), "subjects": {}}
            order[cat] = {}
        if subj not in struct[cat]["subjects"]:
            struct[cat]["subjects"][subj] = {
                "name": SUBJ_NAMES.get(subj, subj),
                "chapters": []   # 数组格式，与 data.js 原格式一致
            }
            order[cat][subj] = []
        # 找到章节在数组中的位置，不存在则追加
        ch_list = struct[cat]["subjects"][subj]["chapters"]
        found = None
        for i, entry in enumerate(ch_list):
            if entry["title"] == ch:
                found = i
                break
        if found is None:
            ch_list.append({"title": ch, "cards": []})
            found = len(ch_list) - 1
        card = {"question": d["question"], "answer": d["answer"]}
        if d.get("tagText"):
            card["tagText"] = d["tagText"]
        diff = d.get("difficulty", "").strip()
        if diff and diff != "0":
            try: card["difficulty"] = int(float(diff))
            except: pass
        if d.get("analysis"):
            card["analysis"] = d["analysis"]
        if d.get("expansion"):
            card["expansion"] = d["expansion"]
        ch_list[found]["cards"].append(card)
    return struct

# ── 生成 JS 文件内容 ──────────────────────────────
def to_js(struct):
    """
    json.dumps 的输出是标准 JSON，
    同时也是合法的 JS 对象字面量（key 有双引号、字符串用双引号）。
    只需要包一层 const cardData = ... ; 即可。
    """
    json_str = json.dumps(struct, indent=4, ensure_ascii=False)
    lines = []
    lines.append("// ==================== 数据配置区 ====================")
    lines.append("// 学科结构：每个学科包含 subjects，每个学科包含 chapters，每个章节包含 cards")
    lines.append("const cardData = " + json_str + ";")
    lines.append("")
    lines.append("// ==================== Node 环境兼容 ====================")
    lines.append("if (typeof module !== 'undefined' && module.exports) {")
    lines.append("    module.exports = cardData;")
    lines.append("}")
    return "\n".join(lines)

# ── 主流程 ─────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("用法：python3 excel_to_js.py <Excel文件> [data.js路径]")
        sys.exit(1)

    excel_path = sys.argv[1]
    js_path = sys.argv[2] if len(sys.argv) > 2 else "data.js"

    if not os.path.exists(excel_path):
        print(f"❌ Excel 文件不存在：{excel_path}")
        sys.exit(1)

    print(f"📖 读取 Excel：{excel_path}")
    data = load_excel(excel_path)
    print(f"✅ 读取到 {len(data)} 行数据")

    print("🧱 整理数据结构...")
    struct = build(data)

    if os.path.exists(js_path):
        bak = js_path + ".bak"
        shutil.copy2(js_path, bak)
        print(f"✅ 已备份原文件：{bak}")

    print("📝 生成 data.js ...")
    js_content = to_js(struct)

    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js_content)

    total = 0
    for c in struct.values():
        for s in c["subjects"].values():
            for ch in s["chapters"]:   # chapters 是数组
                total += len(ch["cards"])

    print(f"✅ 完成！共 {total} 张卡片")
    print(f"📁 输出文件：{os.path.abspath(js_path)}")
    print("\n提示：刷新浏览器即可看到更新后的卡片数据。")

if __name__ == "__main__":
    main()
