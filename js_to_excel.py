#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
【导出】data.js → Excel
用法：python3 js_to_excel.py [data.js路径] [输出.xlsx]

依赖：openpyxl（已内置在 WorkBuddy 环境）
原理：调用 Node.js 执行 data.js，得到 JSON，再写入 Excel
"""

import sys, os, subprocess, json, re, tempfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ 请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)


def js_to_json(js_path):
    """用 Node 执行 data.js，导出 JSON"""
    tmp = tempfile.mktemp(suffix=".js")
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(open(js_path, "r", encoding="utf-8").read())
        f.write("\nprocess.stdout.write(JSON.stringify(cardData));\n")
    try:
        r = subprocess.run(
            ["node", tmp],
            capture_output=True, text=True, timeout=15
        )
        if r.returncode != 0:
            print(f"❌ Node 执行失败：{r.stderr[:300]}")
            return None
        return json.loads(r.stdout)
    except FileNotFoundError:
        print("❌ 未找到 node 命令，请先安装 Node.js")
        return None
    except Exception as e:
        print(f"❌ 执行异常：{e}")
        return None
    finally:
        try: os.unlink(tmp)
        except: pass


def write_excel(struct, out_path):
    wb = openpyxl.Workbook()
    
    # Sheet1: 知识卡片数据
    ws = wb.active
    ws.title = "知识卡片"


    headers = ["学科大类","学科编号","章节名称",
               "问题（正面）","答案（背面）",
               "标签","难度","考点提示","知识拓展"]
    fill_h = PatternFill("solid", fgColor="1A56DB")
    font_h  = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal="center", vertical="center")

    row_i = 2
    for cat_key, cat_obj in struct.items():
        for subj_key, subj_obj in (cat_obj.get("subjects") or {}).items():
            for ch in (subj_obj.get("chapters") or []):
                ch_title = ch.get("title","")
                for card in (ch.get("cards") or []):
                    ws.cell(row=row_i, column=1, value=cat_key)
                    ws.cell(row=row_i, column=2, value=subj_key)
                    ws.cell(row=row_i, column=3, value=ch_title)
                    ws.cell(row=row_i, column=4, value=card.get("question",""))
                    ws.cell(row=row_i, column=5, value=card.get("answer",""))
                    ws.cell(row=row_i, column=6, value=card.get("tagText",""))
                    d = card.get("difficulty","")
                    ws.cell(row=row_i, column=7, value=d if d else "")
                    ws.cell(row=row_i, column=8, value=card.get("analysis",""))
                    ws.cell(row=row_i, column=9, value=card.get("expansion",""))
                    row_i += 1

    widths = [12, 14, 20, 50, 45, 12, 10, 45, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    
    # Sheet2: 使用说明
    ws2 = wb.create_sheet("使用说明")
    instructions = [
        ("字段", "说明", "示例"),
        ("", "", ""),
        ("学科大类", "学科分类（请勿修改）", "law / politics / mazhe 等"),
        ("学科编号", "学科编号（请勿修改）", "xingfa / sixiang 等"),
        ("章节名称", "章节名称（请勿修改）", "刑法概说"),
        ("", "", ""),
        ("问题（正面）", "卡片正面内容，填空用（　）表示", "刑法的基本原则包括（　）、（　）。"),
        ("答案（背面）", "卡片背面内容", "罪刑法定原则、刑法面前人人平等原则"),
        ("", "", ""),
        ("标签", "左上角小标签（选填）", "考点 / 核心考点"),
        ("难度", "难度 1-5（选填）", "1 到 5 的整数"),
        ("考点提示", "考点提示/答题技巧（选填）", "🔥极高频！注意区分XXX和YYY"),
        ("知识拓展", "知识拓展（选填）", "相关法条见刑法第X条"),
        ("", "", ""),
        ("⚠️ 注意事项", "", ""),
        ("1. 括号空格", "填空括号必须用全角空格（　），不能用半角（ ）", "（　）正确，（ ）错误"),
        ("2. 新增章节", "如果章节名称在数据中不存在，脚本会自动新建", ""),
        ("3. 不要删除行", "如果某行不想要了，整行删除即可", ""),
        ("4. 导入命令", "补充完后运行：python3 excel_to_js.py 文件名.xlsx data.js", ""),
    ]
    for r, (col1, col2, col3) in enumerate(instructions, 1):
        ws2.cell(row=r, column=1, value=col1)
        ws2.cell(row=r, column=2, value=col2)
        ws2.cell(row=r, column=3, value=col3)
        if r == 1:
            for c in range(1, 4):
                cell = ws2.cell(row=r, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1A56DB")
                cell.alignment = Alignment(horizontal="center")
        if r >= 15:
            for c in range(1, 4):
                cell = ws2.cell(row=r, column=c)
                cell.font = Font(color="CC0000")
    
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 40
    
    wb.save(out_path)
    return row_i - 2


def main():
    js_path  = sys.argv[1] if len(sys.argv) > 1 else "data.js"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "知识卡片_导出.xlsx"
    if not os.path.exists(js_path):
        print(f"❌ 文件不存在：{js_path}"); sys.exit(1)

    print(f"📖 解析 {js_path} ...")
    struct = js_to_json(js_path)
    if not struct:
        sys.exit(1)

    total = 0
    for c in struct.values():
        for s in (c.get("subjects") or {}).values():
            for ch in (s.get("chapters") or []):
                total += len(ch.get("cards") or [])
    print(f"✅ 解析到 {total} 张卡片")

    print(f"📝 写入 Excel：{out_path}")
    n = write_excel(struct, out_path)
    print(f"✅ 完成！共 {n} 行")
    print(f"📁 文件路径：{os.path.abspath(out_path)}")
    print("\n下一步：在 Excel 里补充新卡片，然后运行导入脚本：")
    print(f"  python3 excel_to_js.py \"{out_path}\" \"{os.path.abspath(js_path)}\"")

if __name__ == "__main__":
    main()
